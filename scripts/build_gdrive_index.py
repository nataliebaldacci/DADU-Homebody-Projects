#!/usr/bin/env python3
"""
Transform gdrive_pdf_mapping.json into an APN-keyed document index
with Google Drive URLs for the Castlehold document portal.

Sources:
1. gdrive_pdf_mapping.json — maps uploaded filenames to Google Drive file IDs
2. Covenants_Final_With_APN_PDF.csv — 3,873 covenants with APN + instrument
3. Parcels_Covenants_with_Buildings spreadsheet — 43,710 parcels with covenant URLs
   (instrument numbers extracted from URLs to match additional covenants)
"""

import json
import re
import csv
from pathlib import Path
from collections import defaultdict

# Paths — corrected to actual locations
GDRIVE_MAPPING = Path("/Users/nataliebaldacci/DADU-Homebody-Projects/DADU/MASTER_ADU_DATA/gdrive_pdf_mapping.json")
COVENANTS_CSV = Path("/Users/nataliebaldacci/DADU-Homebody-Projects/DADU/MASTER_ADU_DATA/Covenants_Final_With_APN_PDF.csv")
COVENANTS_XLSX = Path("/Users/nataliebaldacci/DADU-Homebody-Projects/DADU/MASTER_ADU_DATA/Parcels_Covenants_with_Buildings_20260118_005042.xlsx")
PERMIT_MANIFEST = Path("/Users/nataliebaldacci/DADU-Homebody-Projects/DADU/FINAL_FINAL/Permit_PDFs_Renamed/_manifest.csv")
OUTPUT_FILE = Path("/Users/nataliebaldacci/DADU-Homebody-Projects/data/gdrive_docs_index.json")


def extract_apn_from_filename(filename):
    """Extract APN from filename patterns like PropertyCard_12500000600_139023.pdf"""
    patterns = [
        r'PropertyCard_(\d{11})_\d+\.pdf',
        r'PropertyCard_(\d{9,11})_\d+\.pdf',
        r'Assessor_(\d{11})_\d+\.pdf',
        r'Assessor_(\d{9,11})_\d+\.pdf',
        r'(\d{11})_PropertyCard\.pdf',
        r'(\d{11})_Assessor\.pdf',
        r'(\d{11})_Covenant.*\.pdf',
        r'(\d{11})_Permit.*\.pdf',
        r'(\d{11})_.*\.pdf',
    ]

    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            apn = match.group(1)
            return normalize_apn(apn)
    return None


def extract_instrument_from_url(url):
    """Extract instrument number from davidsonportal covenant URL.
    URL format: .../img/tn/davidson/YYYY/MMDD/INSTRUMENTNUMBER.tif
    """
    if not url:
        return None
    match = re.search(r'/(\d{15,})\.tif', url)
    if match:
        return match.group(1)
    return None


def determine_doc_type(filepath):
    """Determine document type from filepath"""
    filepath_lower = filepath.lower()

    if 'propertycard' in filepath_lower or 'property_card' in filepath_lower:
        return 'property_card'
    elif 'assessor' in filepath_lower:
        return 'assessor_card'
    elif 'covenant' in filepath_lower or 'restrictive' in filepath_lower:
        return 'covenant'
    elif 'permit' in filepath_lower:
        return 'permit'
    elif 'aerial' in filepath_lower or 'screenshot' in filepath_lower:
        return 'aerial'
    elif 'site_plan' in filepath_lower or 'siteplan' in filepath_lower:
        return 'site_plan'
    else:
        return 'other'


def normalize_apn(apn):
    """Normalize APN to 11 digits"""
    if not apn:
        return None
    cleaned = re.sub(r'[^0-9]', '', str(apn))
    if not cleaned or len(cleaned) < 5:
        return None
    if len(cleaned) <= 11:
        return cleaned.zfill(11)
    return cleaned[:11]


def main():
    print(f"Loading Google Drive mapping from {GDRIVE_MAPPING}")

    with open(GDRIVE_MAPPING) as f:
        gdrive_mapping = json.load(f)

    print(f"Found {len(gdrive_mapping)} documents in mapping")

    # Build instrument -> gdrive mapping for covenants
    instrument_to_gdrive = {}
    for filepath, gdrive_info in gdrive_mapping.items():
        if 'Restrictive_Covenants' in filepath or 'Covenant' in filepath:
            filename = Path(filepath).name
            instrument = filename.replace('.pdf', '').replace('.PDF', '')
            instrument_to_gdrive[instrument] = gdrive_info

    print(f"Found {len(instrument_to_gdrive)} covenant documents in Drive")

    # Build permit_number -> gdrive mapping
    permit_to_gdrive = {}
    for filepath, gdrive_info in gdrive_mapping.items():
        if 'Permit_PDFs' in filepath:
            filename = Path(filepath).name
            # Original permit PDFs have UUID filenames
            permit_to_gdrive[filename] = gdrive_info

    print(f"Found {len(permit_to_gdrive)} permit documents in Drive")

    # Build APN-keyed index
    apn_docs = defaultdict(lambda: {
        'property_cards': [],
        'assessor_cards': [],
        'covenants': [],
        'permits': [],
        'aerials': [],
        'site_plans': [],
        'other': []
    })

    matched = 0
    unmatched = 0

    # ── Process property cards and other documents ──
    for filepath, gdrive_info in gdrive_mapping.items():
        filename = Path(filepath).name
        apn = extract_apn_from_filename(filename)

        if apn:
            doc_type = determine_doc_type(filepath)

            # Skip covenants here — handled from CSV/XLSX below
            if doc_type == 'covenant':
                continue

            # Skip permits here — handled from manifest below
            if doc_type == 'permit':
                continue

            doc_entry = {
                'filename': filename,
                'gdrive_url': gdrive_info['url'],
                'download_url': gdrive_info['download_url'],
                'file_id': gdrive_info['file_id']
            }

            if doc_type == 'property_card':
                apn_docs[apn]['property_cards'].append(doc_entry)
            elif doc_type == 'assessor_card':
                apn_docs[apn]['assessor_cards'].append(doc_entry)
            elif doc_type == 'aerial':
                apn_docs[apn]['aerials'].append(doc_entry)
            elif doc_type == 'site_plan':
                apn_docs[apn]['site_plans'].append(doc_entry)
            else:
                apn_docs[apn]['other'].append(doc_entry)

            matched += 1
        else:
            unmatched += 1

    print(f"Matched {matched} property/aerial docs to APNs ({unmatched} unmatched)")

    # ── Process covenants: Source 1 — Covenants_Final_With_APN_PDF.csv ──
    covenant_count = 0
    seen_covenant_keys = set()  # (apn, instrument) to deduplicate

    if COVENANTS_CSV.exists():
        print(f"\nLoading covenant mapping from {COVENANTS_CSV}")
        with open(COVENANTS_CSV, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                apn = normalize_apn(row.get('APN', ''))
                instrument = row.get('Instrument_Clean', '').strip()
                has_pdf = row.get('has_pdf', '').lower() == 'true'

                if apn and instrument and has_pdf:
                    key = (apn, instrument)
                    if key in seen_covenant_keys:
                        continue
                    if instrument in instrument_to_gdrive:
                        gdrive_info = instrument_to_gdrive[instrument]
                        doc_entry = {
                            'filename': f"{instrument}.pdf",
                            'instrument': instrument,
                            'rec_date': row.get('Rec. Date', ''),
                            'grantor': row.get('Grantor', ''),
                            'gdrive_url': gdrive_info['url'],
                            'download_url': gdrive_info['download_url'],
                            'file_id': gdrive_info['file_id']
                        }
                        apn_docs[apn]['covenants'].append(doc_entry)
                        seen_covenant_keys.add(key)
                        covenant_count += 1

        print(f"  Source 1 (CSV): Matched {covenant_count} covenants")

    # ── Process covenants: Source 2 — 43K spreadsheet with instrument URLs ──
    covenant_xlsx_count = 0
    if COVENANTS_XLSX.exists():
        try:
            import openpyxl
            print(f"\nLoading comprehensive covenant mapping from {COVENANTS_XLSX}")
            wb = openpyxl.load_workbook(COVENANTS_XLSX, read_only=True)
            ws = wb.active

            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            apn_idx = headers.index('APN') if 'APN' in headers else None
            cov_url_idx = headers.index('Restrictive Covenant') if 'Restrictive Covenant' in headers else None
            date_idx = headers.index('Date CR Recorded') if 'Date CR Recorded' in headers else None

            if apn_idx is not None and cov_url_idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    apn_raw = row[apn_idx]
                    cov_url = row[cov_url_idx]

                    apn = normalize_apn(apn_raw)
                    instrument = extract_instrument_from_url(str(cov_url) if cov_url else '')

                    if apn and instrument:
                        key = (apn, instrument)
                        if key in seen_covenant_keys:
                            continue
                        if instrument in instrument_to_gdrive:
                            gdrive_info = instrument_to_gdrive[instrument]
                            rec_date = ''
                            if date_idx is not None and row[date_idx]:
                                try:
                                    rec_date = row[date_idx].strftime('%Y-%m-%d')
                                except:
                                    rec_date = str(row[date_idx])

                            doc_entry = {
                                'filename': f"{instrument}.pdf",
                                'instrument': instrument,
                                'rec_date': rec_date,
                                'grantor': '',
                                'gdrive_url': gdrive_info['url'],
                                'download_url': gdrive_info['download_url'],
                                'file_id': gdrive_info['file_id']
                            }
                            apn_docs[apn]['covenants'].append(doc_entry)
                            seen_covenant_keys.add(key)
                            covenant_xlsx_count += 1

            wb.close()
            print(f"  Source 2 (XLSX): Matched {covenant_xlsx_count} additional covenants")

        except ImportError:
            print("  openpyxl not installed, skipping XLSX source")
        except Exception as e:
            print(f"  Error reading XLSX: {e}")

    total_covenants = covenant_count + covenant_xlsx_count
    print(f"  Total covenants matched: {total_covenants}")

    # ── Process permits from manifest ──
    permit_count = 0
    if PERMIT_MANIFEST.exists():
        print(f"\nLoading permit manifest from {PERMIT_MANIFEST}")
        with open(PERMIT_MANIFEST, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                apn = normalize_apn(row.get('APN', ''))
                permit_number = row.get('permit_number', '').strip()
                original_filename = row.get('original_filename', '').strip()

                if apn and original_filename and original_filename in permit_to_gdrive:
                    gdrive_info = permit_to_gdrive[original_filename]
                    doc_entry = {
                        'filename': row.get('renamed_filename', original_filename),
                        'permit_number': permit_number,
                        'address': row.get('address', ''),
                        'gdrive_url': gdrive_info['url'],
                        'download_url': gdrive_info['download_url'],
                        'file_id': gdrive_info['file_id']
                    }
                    apn_docs[apn]['permits'].append(doc_entry)
                    permit_count += 1

        print(f"  Matched {permit_count} permits to APNs with Drive links")

    total_docs = matched + total_covenants + permit_count
    print(f"\nTotal parcels with documents: {len(apn_docs)}")
    print(f"Total documents indexed: {total_docs}")

    # Build output structure
    output = {
        '_schema_version': '2.0',
        '_description': 'Google Drive document index keyed by APN for Castlehold document portal',
        '_generated': __import__('datetime').datetime.now().isoformat(),
        '_total_parcels': len(apn_docs),
        '_total_documents': total_docs,
        'parcels': dict(apn_docs)
    }

    # Count by type
    type_counts = defaultdict(int)
    for apn, docs in apn_docs.items():
        for doc_type, doc_list in docs.items():
            type_counts[doc_type] += len(doc_list)

    output['_document_counts'] = dict(type_counts)

    print(f"\nDocument counts by type:")
    for doc_type, count in sorted(type_counts.items()):
        if count > 0:
            print(f"  {doc_type}: {count}")

    # Write output
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\nWrote {OUTPUT_FILE}")
    print(f"File size: {OUTPUT_FILE.stat().st_size / 1024 / 1024:.2f} MB")

    # Report gaps
    print(f"\n=== GAPS (files on Drive without APN mapping) ===")
    print(f"  Covenants on Drive: {len(instrument_to_gdrive)}")
    print(f"  Covenants indexed: {total_covenants}")
    print(f"  Covenant gap: {len(instrument_to_gdrive) - total_covenants}")


if __name__ == '__main__':
    main()
