from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "All Live Links"

BASE = "https://nataliebaldacci.github.io/DADU-Homebody-Projects/"

# Styles
header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="3A5566")
section_font = Font(name="Arial", bold=True, size=11, color="3A5566")
section_fill = PatternFill("solid", fgColor="F0EBE1")
link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
body_font = Font(name="Arial", size=10)
body_font_bold = Font(name="Arial", size=10, bold=True)
wheat_fill = PatternFill("solid", fgColor="CBB279")
wheat_font = Font(name="Arial", bold=True, size=11, color="3A5566")
thin_border = Border(
    bottom=Side(style="thin", color="E2E2E0")
)

# Column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 75
ws.column_dimensions["E"].width = 18

# Headers
headers = ["#", "Page Name", "Nav Location", "Full URL", "Status"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

row = 2
num = 1

def add_section(sheet, r, title):
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = sheet.cell(row=r, column=1, value=title)
    c.font = wheat_font
    c.fill = wheat_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[r].height = 24
    return r + 1

def add_subsection(sheet, r, title):
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = sheet.cell(row=r, column=1, value=title)
    c.font = section_font
    c.fill = section_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[r].height = 22
    return r + 1

def add_link(sheet, r, n, name, nav, filename, status="Live"):
    sheet.cell(row=r, column=1, value=n).font = body_font
    sheet.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    sheet.cell(row=r, column=2, value=name).font = body_font_bold
    sheet.cell(row=r, column=3, value=nav).font = body_font
    url = BASE + filename
    c = sheet.cell(row=r, column=4, value=url)
    c.font = link_font
    c.hyperlink = url
    sheet.cell(row=r, column=5, value=status).font = body_font
    sheet.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    for col in range(1, 6):
        sheet.cell(row=r, column=col).border = thin_border
    return r + 1, n + 1

# ── HOMEPAGE + ENTRY POINTS ──
row = add_section(ws, row, "HOMEPAGE + ENTRY POINTS")
row, num = add_link(ws, row, num, "Homepage", "Logo click", "index.html")
row, num = add_link(ws, row, num, "Am I Eligible?", "Get Started CTA", "am_i_eligible.html")
row, num = add_link(ws, row, num, "Property Search", "Search icon", "property_search.html")
row, num = add_link(ws, row, num, "Pricing", "PRICING nav link", "homebody_dadu_pricing.html")

# ── EXPLORE > LEARN ──
row = add_section(ws, row, "EXPLORE")
row = add_subsection(ws, row, "Learn")
row, num = add_link(ws, row, num, "What is a DADU?", "Explore > Learn", "what_is_dadu.html")
row, num = add_link(ws, row, num, "History & Timeline", "Explore > Learn", "dadu_history.html")
row, num = add_link(ws, row, num, "Requirements", "Explore > Learn", "dadu_building_requirements.html")
row, num = add_link(ws, row, num, "Zoning Standards", "Explore > Learn", "dadu_zoning_standards.html")
row, num = add_link(ws, row, num, "Code & Legislation", "Explore > Learn", "dadu_code_legislation_v3.html")
row, num = add_link(ws, row, num, "Permit Process", "Explore > Learn", "permit_process_timeline.html")

row = add_subsection(ws, row, "Discover")
row, num = add_link(ws, row, num, "Eligibility Map", "Explore > Discover", "dadu_eligibility_map.html")
row, num = add_link(ws, row, num, "Property Search", "Explore > Discover", "property_search.html")
row, num = add_link(ws, row, num, "DADUs Near Me", "Explore > Discover", "dadu_near_me_v2.html")
row, num = add_link(ws, row, num, "Opportunity Explorer", "Explore > Discover", "dadu_opportunity_explorer_v2.html")

row = add_subsection(ws, row, "User Types")
row, num = add_link(ws, row, num, "Homeowner", "Explore > User Types", "user-homeowners.html")
row, num = add_link(ws, row, num, "Contractor", "Explore > User Types", "contractor_marketplace.html")
row, num = add_link(ws, row, num, "Designer / Architect", "Explore > User Types", "designer_resources.html")
row, num = add_link(ws, row, num, "Municipal / Agency", "Explore > User Types", "municipal_dashboard.html")
row, num = add_link(ws, row, num, "Legal / Appraiser", "Explore > User Types", "legal_resources.html")

# ── BUILD ──
row = add_section(ws, row, "BUILD")
row = add_subsection(ws, row, "Plan")
row, num = add_link(ws, row, num, "Project Planner", "Build > Plan", "project_planner.html")
row, num = add_link(ws, row, num, "Interactive Checklist", "Build > Plan", "project_checklist.html")
row, num = add_link(ws, row, num, "Draw DADU on Parcel", "Build > Plan", "draw_dadu_on_parcel.html")

row = add_subsection(ws, row, "Design & Calculate")
row, num = add_link(ws, row, num, "Site Plan Finder", "Build > Design & Calculate", "site_plan_downloads.html")
row, num = add_link(ws, row, num, "Cost Estimator", "Build > Design & Calculate", "project_cost_estimator.html")
row, num = add_link(ws, row, num, "ROI Calculator", "Build > Design & Calculate", "roi_calculator.html")
row, num = add_link(ws, row, num, "Size Calculator", "Build > Design & Calculate", "size_calculator.html")
row, num = add_link(ws, row, num, "Tax Calculator", "Build > Design & Calculate", "property_tax_calculator.html")

row = add_subsection(ws, row, "Hire")
row, num = add_link(ws, row, num, "Contractor Marketplace", "Build > Hire", "contractor_marketplace.html")

row = add_subsection(ws, row, "File")
row, num = add_link(ws, row, num, "Determine Forms", "Build > File", "determine_forms_required.html")
row, num = add_link(ws, row, num, "Form Filler", "Build > File", "legal_form_filler.html")
row, num = add_link(ws, row, num, "Owner Occupancy", "Build > File", "owner_occupancy.html")
row, num = add_link(ws, row, num, "STR Permit", "Build > File", "str_permit.html")

# ── DATA ──
row = add_section(ws, row, "DATA")
row = add_subsection(ws, row, "Activity")
row, num = add_link(ws, row, num, "Permit Dashboard", "Data > Activity", "permit_activity_dashboard.html")
row, num = add_link(ws, row, num, "Contractor Dashboard", "Data > Activity", "contractor_dashboard.html")
row, num = add_link(ws, row, num, "Market Trends", "Data > Activity", "market_trends.html")

row = add_subsection(ws, row, "Reports")
row, num = add_link(ws, row, num, "Eligibility Report", "Data > Reports", "eligibility_report.html")
row, num = add_link(ws, row, num, "Property Intelligence", "Data > Reports", "property-report-card.html")
row, num = add_link(ws, row, num, "Project Report", "Data > Reports", "project_report.html")
row, num = add_link(ws, row, num, "Contractor Report", "Data > Reports", "dadu_reports_store.html")
row, num = add_link(ws, row, num, "Market Analysis", "Data > Reports", "dadu_reports_store.html")
row, num = add_link(ws, row, num, "Area Analysis", "Data > Reports", "dadu_reports_store.html")

row = add_subsection(ws, row, "Records")
row, num = add_link(ws, row, num, "Permit Explorer", "Data > Records", "nashville_permit_explorer_v3.html")
row, num = add_link(ws, row, num, "Permit Site Plans", "Data > Records", "site_plan_downloads.html")
row, num = add_link(ws, row, num, "Recorded Documents", "Data > Records", "dadu_documents_portal.html")
row, num = add_link(ws, row, num, "Restrictive Covenants", "Data > Records", "restrictive_covenants_v2.html")
row, num = add_link(ws, row, num, "Zoning Documents", "Data > Records", "overlay-districts.html")
row, num = add_link(ws, row, num, "PDF Database", "Data > Records", "pdf_database_lookup.html")

# ── NON-CANONICAL ──
row = add_section(ws, row, "NON-CANONICAL PAGES (not in nav, publicly accessible)")

non_canonical = [
    ("About Platform Infographic", "about_platform_infographic.html"),
    ("ADU Permit Map", "adu_permit_map.html"),
    ("Am I Eligible (Compact)", "am_i_eligible_compact.html"),
    ("Area Analysis Report", "area_analysis_report.html"),
    ("Castlehold Homepage (Flat)", "castlehold_homepage_flat.html"),
    ("Contractor Advertising", "contractor_advertising.html"),
    ("Contractor Portal (Old)", "contractor_portal.html"),
    ("Contractor Report", "contractor_report.html"),
    ("DADU Build Explorer v1", "dadu_build_explorer.html"),
    ("DADU Build Explorer v2", "dadu_build_explorer_v2.html"),
    ("DADU Build Tool", "dadu_build_tool.html"),
    ("DADU Build Tool Backup", "dadu_build_tool_backup.html"),
    ("Code & Legislation v4", "dadu_code_legislation_v4.html"),
    ("Code & Legislation v5", "dadu_code_legislation_v5.html"),
    ("Contractors Infographic", "dadu_contractors_infographic.html"),
    ("Design Standards", "dadu_design_standards.html"),
    ("Eligibility Checklist (Old)", "dadu_eligibility_checklist.html"),
    ("Eligibility Flowchart (Old)", "dadu_eligibility_flowchart.html"),
    ("Explorer ATTOM v1", "dadu_explorer_attom.html"),
    ("Explorer ATTOM v2", "dadu_explorer_attom_v2.html"),
    ("Explorer v2 (Old)", "dadu_explorer_v2.html"),
    ("Near Me Locator (Old)", "dadu_near_me_locator.html"),
    ("Near Me v3", "dadu_near_me_v3.html"),
    ("DADU Platform (Old)", "dadu_platform.html"),
    ("DADU Property Report (Old)", "dadu_property_report.html"),
    ("Property Viewer v3", "dadu_property_viewer_v3.html"),
    ("Report Connected", "dadu_report_connected.html"),
    ("Report Full", "dadu_report_full.html"),
    ("DADU Resources (Old)", "dadu_resources.html"),
    ("Symbium Map POC", "dadu_symbium_map.html"),
    ("Designer Portal (Old)", "designer_portal.html"),
    ("Feature: Documents", "feature-documents.html"),
    ("Feature: Eligibility Map", "feature-eligibility-map.html"),
    ("Feature: Property Search", "feature-property-search.html"),
    ("Features Overview", "features.html"),
    ("Footprints POC", "footprints_proof_of_concept.html"),
    ("Header Template", "homebody_header.html"),
    ("Homepage v1 (Old)", "homebody_index.html"),
    ("Homepage Map (Old)", "homebody_index_map.html"),
    ("Homepage v3 (Old)", "homebody_index_v3.html"),
    ("Homepage v4 (Old)", "homebody_index_v4.html"),
    ("Homebody Main (Old)", "homebody_main.html"),
    ("Homeowner Portal (Old)", "homeowner_portal.html"),
    ("Index PQ (Alt)", "index_pq.html"),
    ("Permit Analytics (Old)", "nashville_permit_analytics.html"),
    ("Parcel Footprint Map", "parcel_footprint_map.html"),
    ("Permit Explorer (Old)", "permit_explorer.html"),
    ("Property Intelligence Report", "property_intelligence_report.html"),
    ("Property Report (Old)", "property_report.html"),
    ("Short Term Rental Permit (Old)", "short_term_rental_permit.html"),
    ("Trade Permits", "trade_permits.html"),
    ("User Types (Old)", "user-types.html"),
    ("Zoning Documents (Old)", "zoning_documents.html"),
]

for name, filename in non_canonical:
    row, num = add_link(ws, row, num, name, "Not in nav", filename, "Non-canonical")

# ── SAMPLE REPORTS ──
row = add_section(ws, row, "SAMPLE REPORTS")

samples = [
    ("Sample Comparables Report", "sample_reports/sample_comparables_report.html"),
    ("Sample Contractor Report", "sample_reports/sample_contractor_report.html"),
    ("Sample Cost Estimate Report", "sample_reports/sample_cost_estimate_report.html"),
    ("Sample Covenant Report", "sample_reports/sample_covenant_report.html"),
    ("Sample Eligibility Report", "sample_reports/sample_eligibility_report.html"),
    ("Sample Market Stats Report", "sample_reports/sample_market_stats_report.html"),
    ("Sample Neighbors Report", "sample_reports/sample_neighbors_report.html"),
    ("Sample Permit History Report", "sample_reports/sample_permit_history_report.html"),
    ("Sample Property Report", "sample_reports/sample_property_report.html"),
    ("Sample Zoning Report", "sample_reports/sample_zoning_report.html"),
]

for name, filename in samples:
    row, num = add_link(ws, row, num, name, "Sample report", filename, "Sample")

# ── SUMMARY ROW ──
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
c = ws.cell(row=row, column=1, value=f"TOTAL: {num - 1} live HTML files  |  43 nav items (39 unique pages)  |  53 non-canonical  |  10 sample reports")
c.font = Font(name="Arial", bold=True, size=10, color="3A5566")
c.alignment = Alignment(horizontal="center")

out = "/sessions/amazing-optimistic-fermi/mnt/DADU-Homebody-Projects/Homebody_Projects_All_Live_Links.xlsx"
wb.save(out)
print(f"Saved to {out}")
